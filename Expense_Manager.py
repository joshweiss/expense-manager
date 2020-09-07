import PySimpleGUI as sg
import openpyxl
import matplotlib.pyplot as plt

expenses_excel = openpyxl.load_workbook("Expenses.xlsx")

show_monthly_totals = False
charges = []
categories = ['Food', 'Bills', 'Fun', 'Baby', 'One-Time', 'Tzedaka', 'Shopping', 'Received', 'Savings']
sums = {}
for cat in categories:       # Sets the Sums of each category to begin at 0
    sums[cat] = 0


#Setting up the Gui
sg.ChangeLookAndFeel('BlueMono')

layout = [
        [sg.Text('Month'), sg.Combo(['October', 'November', 'December', 'January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September'], key = 'month')],
        [sg.Text('What is the charge?'),sg.InputText(key = 'detail', size = (30,1), do_not_clear=False)],
        [sg.Text('Cost?',size=(19,1)),sg.InputText(key = 'cost', size = (10,1), do_not_clear=False)],
        [sg.Text('Type of Expense'), sg.Combo(categories, key = 'category')],
        [sg.Submit(),sg.Exit(), sg.Button('Show Monthly Totals')]
        ]

window  = sg.Window("Josh\'s Expense Manager").Layout(layout)

while True:      
    button, values = window.Read()
    if button is None or button == 'Exit':
        break
    #Adding the values from the gui into a list of dictionaries for each time the form was submitted
    if button == 'Submit':
        charges.append({'detail': values['detail'], 'cost': float(values['cost']), 'category': values['category'], 'month': values['month']})
    if button == 'Show Monthly Totals':
        month_for_show = values['month']
        show_monthly_totals = True
        break
            
window.Close()

#Filling up the excel sheet
for i in range(0,len(charges)):                                      #running through the list of submitted forms
    monthly_sheet = charges[i]['month'] + ' Expenses'                
    
    if monthly_sheet in expenses_excel.sheetnames:                   #checking if its a new month
        charges_sheet = expenses_excel[monthly_sheet]                #defining the sheets according to the chosen month
        totals_sheet = expenses_excel[charges[i]['month'] + ' Totals']
    else:
        expenses_excel.create_sheet(monthly_sheet)                   # if its a new month creating the new sheet
        charges_sheet = expenses_excel[monthly_sheet]                #defining the expenses sheet and giving it starting headers
        charges_sheet['A1'].value = 'Expense'
        charges_sheet['B1'].value = 'Cost'
        charges_sheet['C1'].value = 'Category'
        
        expenses_excel.create_sheet(charges[i]['month'] + ' Totals')    #creating the totals sheet and giving it starting headers
        totals_sheet = expenses_excel[charges[i]['month'] + ' Totals']
        totals_sheet['A1'].value = 'Category'
        totals_sheet['B1'].value = 'Sum'
        totals_sheet['A2'].value = 'Total Spent'
        totals_sheet['B2'].value = float(0)

        for j, cat in enumerate(categories, start=3):                 #writing the categories with 0 sum to the totals sheet
            totals_sheet.cell(j,1).value = cat
            totals_sheet.cell(j,2).value = 0

        totals_sheet.cell(j-1,2).value = float(1000)                  #starting the savings at 1000 shekels (altshuler)
        totals_sheet.cell(j,1).value = 'Fixed Expenses'               #creating the fixed expenses
        totals_sheet.cell(j,2).value = float(7790)
        totals_sheet.cell(j,3).value = 'Rent = 5090, Car = 1000, Wifi = 90, Vaad Bayit = 295,  Cellcom = 50, Maccabbi Sheli = 80, Arnona/Water = 1185'
 
    row = 2
    while charges_sheet.cell(row,1).value is not None:                #reaching the bottom of the expenses sheet
        row+=1
    
    charges_sheet.cell(row,1).value = charges[i]['detail']            #inputting the expense details to the expense sheet
    charges_sheet.cell(row,2).value = charges[i]['cost']
    charges_sheet.cell(row,3).value = charges[i]['category']
    
    j = 3
    while totals_sheet.cell(j,1).value is not None:                   
        if totals_sheet.cell(j,1).value == charges[i]['category']:    #finding the category of the expense
            totals_sheet.cell(j,2).value += charges[i]['cost']        #adding the expense cost to the total of its category
            if charges[i]['category'] != 'Received':              #don't add the expense to the total if its money received
                totals_sheet['B2'].value += charges[i]['cost']
            break
        j+=1

#Displaying the totals
if show_monthly_totals:
    charges_sheet = expenses_excel[month_for_show + ' Expenses']
    totals_sheet = expenses_excel[month_for_show + ' Totals']
    j = 3
    cat_totals = []
    leftover = 12000
    while totals_sheet.cell(j,1).value is not None:
        if totals_sheet.cell(j,1).value != 'Received':
            cat_totals.append(totals_sheet.cell(j,2).value)
            leftover -= totals_sheet.cell(j,2).value
        j+=1
    expenses_cats = categories
    expenses_cats.remove('Received')
    expenses_cats.append('Leftover')
    cat_totals.append(leftover)
    print(leftover)
    plt.pie(cat_totals, autopct='%1.1f%%')
    plt.legend(labels = expenses_cats)
    plt.show()

    
    """
    sg.ChangeLookAndFeel('GreenTan')
    layout_show = [
                    [sg.Text('Expenses for the month of ' + month_for_show)],
                    [sg.Canvas(size=(100,100))],
                    [sg.Exit()]
                  ]
    
    window  = sg.Window("Josh\'s Expense Manager").Layout(layout_show)
    while True:      
        button, values = window.Read()
        if button is None or button == 'Exit':
            break
        
    window.Close()
    
    """

expenses_excel.save("Expenses.xlsx")
